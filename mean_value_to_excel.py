#! C:\Users\il\AppData\Local\Programs\Python\Python36\python.exe
#
# TODO: clean up, add/correct comments, config.ini file, ...
# tested -and in production- on databases with 1 measurement for each tag (no use of field keys)
# timestamp set to nanosecond
# ... other prerequisites : to add
#
"""
VERSION: 0010
runs on InfluxDB v0.9
This program reads an Excel file -xlsx- and returns values of the tags according the listed periods
The filename is hard coded
the format must be respected:
    ROW 3
        beginning at column D: tags. Tags must have identical name as in InfluxDB
            an empty column means the end of the tag list.
    ROW 4 and following
    column A : begin date-time of one period
    column B : end date-time of one period
        for every row with value column D is empty:
            if no value in A and value B == 'now': the last value for each tag will be written
                Note: because the Excel column A and B are configured as date-time,
                a '0' (i.e. zero)  will be represented as 00/01/1970 0:00 resulting in no values because of invalid date-time
            if no value in A and value B (time End) < now (actual time of server): the last value before time End will be written
            if value A (time Begin) < value B (time End) : the mean time over the period time Begin to End will be written
        if value D is empty: the engine stops and write values to Excel
Return:
    for each row (starting at 4) the sample date-time (start moment of sampling) (column C)
        and the corresponding mean values of the tags over the period mentioned in column A and B
"""
import time
from influxdb import InfluxDBClient
# OPENPYXL
from openpyxl import Workbook, load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
import datetime

startRunTime = datetime.datetime.now()
print('startRunTime: ', startRunTime)
#
#
# """ USE THIS SECTION IN CASE OF ARGUMENT PARSING IN COMMAND LINE"""
# import argparse
# parser = argparse.ArgumentParser()
# # parser.add_argument('fn', type=str, default='actual_value.xlsx',
# #                                      help='Read input and store program output in the file passed')
# parser.add_argument('fileName', type=str, help='Read input and pass the file name to the programm')
# args = parser.parse_args()
# print('args = ', args.fileName)
# xlsxFileName = str(args.fileName) # USE THIS SECTION IN CASE OF ARGUMENT PARSING IN COMMAND LINE

xlsxFileName = "mean_value_to_excel.xlsx"     #"mean_value_to_excel.xlsx"   # USE THIS SECTION IN CASE OF FIXED FILE NAME - NO ARG PARSE
# example of terminal command: C:\...\...>python mean_value_to_excel.py "mean_value_to_excel.xlsx"


# location of data file to ingest is hard coded
# modus = "productionDB_local_code"   # code running on local PC using production influx db
# modus = "test"                    # code running on local PC using local influx db
# modus = "productionDB_localCode_localFiles" # local code, local file, production influx db
# modus = "production-Hist01"            # code running on production server using production influx db
# modus = "production-Hist02"              # code running on production server using production influx db
modus = "hist02_localCode_localFiles"    # code running on local PC using production influx db

try:
    if modus == "test":
        DirIn = "\\\\server1.domain.be\\dir1\\dir2\\input\\"
        DirOut = "\\\\server1.domain.be\\dir1\\dir2\\output\\"

        host = "10.11.12.13"
        port = 8086
        user = "my_root"
        password = "my_root"
        dbname = "my_influxdb"

    elif modus == "production-Hist01":
        """Other set of parameters"""
    elif modus == "productionDB_local_code":
        """Other set of parameters"""
    elif modus == "productionDB_localCode_localFiles":
         """Other set of parameters"""
    elif modus == "production-Hist02" :
        """Other set of parameters"""
    elif modus == "hist02_localCode_localFiles":
        """Other set of parameters"""

    else:
        # *** SEND ERROR MAIL ***
        raise SystemExit('Wrong modus, modify code.')
except:
    # *** SEND ERROR MAIL ***
    raise SystemExit('No modus, modify code.')


xlsxDirFileNameIn = DirOut + xlsxFileName
xlsxDirFileNameOut = DirOut + xlsxFileName
print(xlsxDirFileNameIn)

wb = load_workbook(filename=xlsxDirFileNameIn)
ws = wb['sheet1']
# print("value A1: ", ws['A1'].value)

""" start column tags"""
startColumnTags = 4 # Fourth column  == column D
startColumnTags = startColumnTags - 1 # startColumn starting from 0


def time_ISO8601_to_Influx_epoch(timeString):
    """
    convert time format ISO8601 to epoch time, precision nanoseconds
    :param timeString: format: "2018-02-13T06:44:25.123Z" decimal precision my vary
    (:param precision : 9 (seconds) to maximal 9 (nanoseconds))
    :return: time format Epoch - integer e.g. 1518504265123000000
    """
    timeEpoch = int(time.mktime(time.strptime(timeString[:19], '%Y-%m-%dT%H:%M:%S')))
    # #print("timeEpoch without timezone correction: ", timeEpoch)
    # timeZoneOffset_HHMM = time.strftime("%z", time.gmtime(time.time()))
    # #print("timezoneOffset_HHMM: ",timeZoneOffset_HHMM)
    # timeZoneOffset_s = (int(timeZoneOffset_HHMM[1:3]) * 60 + int(timeZoneOffset_HHMM[3:5])) * 60
    # print("timeZoneOffset_s: ", timeZoneOffset_s)
    # # timeEpoch = timeEpoch + timeZoneOffset_s
    # # print("timeEpoch with timezone correction: ", timeEpoch)
    decimalSeconds = timeString[19:].strip('.Z')
    #print(decimalSeconds)
    decimalSeconds = '{:0<9}'.format(decimalSeconds)
    #print(decimalSeconds)
    timeInflux = str(timeEpoch) + decimalSeconds
    #print("time_ISO8601_to_Influx_epoch(timeString): ", time_ISO8601_to_Influx_epoch(datestring))
    return int(timeInflux)


timeB = '2018-04-17T08:00:00.000' # begin of query run time period
timeE = '2018-04-18T08:00:00.000' # end of query run time period
timeB_Epoch = time_ISO8601_to_Influx_epoch(timeB)
timeE_Epoch = time_ISO8601_to_Influx_epoch(timeE)
# print('mean values from {0} to {1}'.format(timeB, timeE))

diff_time = time_ISO8601_to_Influx_epoch(timeE) - time_ISO8601_to_Influx_epoch(timeB)
# print(diff_time/1000000000)
#************************


def db_lookup(client):
    """
    Give me a client  and i'll return the list of databases
    :param client: host, port, user, password
    :return: list databases: format: [{'name': 'database01'}, {'name': 'database02'}, {'name': 'database03'}]
    """
    dblist_dict= client.get_list_database()
    # print("def db_lookup 010:", dblist_dict)
    # print("def db_lookup 020:", dblist_dict[3]['name'])
    # for element in dblist_dict:
    #     print("db_lookup 3:", element['name'])
    return dblist_dict


def measurements_lookup(client, database):
    """
    Give me a database and i'll return the list of measurements
    :param client: host, port, user, password
    :param database: string of database name, format 'database01'
    :return: list of measurements: format:
    """
    client.switch_database(database)
    mlist_dict = client.get_list_measurements()
    # print("def measurements_lookup 010:", mlist_dict[:10])
    return mlist_dict


def all_measurements_lookup(client):
    """
    Give me a client and i'll return the list of all measurements of all databases of that client
    :param client: host, port, user, password
    :return: list of all measurements: format list of dict
        [{'database07': [{'name': 'TAG1.70.60.50'}, {'name': 'TAG2.70.50.49'}, ....
         {'database01': [{'name': 'TAG1.10.11.12'}, {'name': 'TAG2.10.21.22'}, {'name': 'TAG3.10.31.32'}, ...]
    """
    dbs_dict = db_lookup(client)
    m_list_dict = []
    for db in dbs_dict:
        m_list_dict.append({db['name']: measurements_lookup(client, db['name'])})
    # print("def all_measurements_lookup 1: ", m_list_dict[:10])
    return m_list_dict


def all_measurements_to_list(ms, filename='n'):
    """
    Give me all measurements and i'll return a csv file on the indicated directory/filename
    :param ms:list of all measurements: format list of dict
        [{'database07': [{'name': 'TAG1.70.60.50'}, {'name': 'TAG2.70.50.49'}, ....
         {'database01': [{'name': 'TAG1.10.11.12'}, {'name': 'TAG2.10.21.22'}, {'name': 'TAG3.10.31.32'}, ...]
    :param filename: complete directory and filename
                default filename = 'n' : no file output
    :return: m_list: format ['TAG1.70.60.50,database07', 'TAG2.70.50.49,database07', ...]
            and a file of this list on the location of your choice
    """
    m_list = []
    for el in ms:
        # print("def all_measurements_to_list 010: ", el)
        for k, v in el.items():
            # print("def all_measurements_to_list 020: ", k, v)
            for meas in v:
                m_list.append(k + ',' + meas['name'])
                # print("def all_measurements_to_list 030: ", k + ' : ' + meas['name'])
        # print(m_list)
    # print("def all_measurements_to_list 040: length measurement list: ", len(m_list))
    #
    if filename != 'n':
        mlist = open(filename, 'w')
        for el in m_list:
            mlist.writelines(el + '\n')
        mlist.close()
        # print("def all_measurements_to_list 050: list saved as: ", filename)
    else:
        # print("def all_measurements_to_list 051: list not saved as file.")
        pass
    return m_list


def measurement_database_lookup(all_measurements, measurement, debug='n'):
    """
    Give me all measurements and i'll return all databases where i found your measurement
    can print message for debugging
    :param all_measurements: list of all measurements: format list of dict
        [{'database07': [{'name': 'TAG1.70.60.50'}, {'name': 'TAG2.70.50.49'}, ....
         {'database01': [{'name': 'TAG1.10.11.12'}, {'name': 'TAG2.10.21.22'}, {'name': 'TAG3.10.31.32'}, ...]
    :param measurement: a single measurement in format string : 'TAG1.10.11.12'
    :param debug = 'y' : print messages
    :return: dictionary in format: {'measurement': 'TAG1.10.11.22', 'database': ['database01', 'database02']}
    """
    resultList = []
    db_list = []
    result_db_dict = {'measurement': measurement, 'database': db_list}

    for el in all_measurements:
        if debug == 'y' : print("def measurement_database_lookup 010: ", el)
        for k, v in el.items():
            if debug == 'y': print("def measurement_database_lookup 020: ", k, v)
            for meas in v:
                if meas['name'] == measurement:
                    resultList.append({'database': k, 'measurement': measurement})
                    # print("def measurement_database_lookup 030: ", k + ' : ' + measurement)
                    db_list.append(k)
    if debug == 'y': print("def measurement_database_lookup 040: ", resultList)
    if debug == 'y': print("def measurement_database_lookup 041: ", db_list)
    if debug == 'y': print("def measurement_database_lookup 042: ", result_db_dict)
    if len(result_db_dict['database']) == 0:
        message = 'not in any database'
    elif len(result_db_dict['database']) == 1:
        message = 'in just 1 database'
    elif len(result_db_dict['database']) > 1:
        message = f'in {len(result_db_dict["database"])} databases'
    if debug == 'y': print('def measurement_database_lookup 050: ', f'Measurement "{result_db_dict["measurement"]}" {message} on host "{host}".')
    return result_db_dict



#************************
client = InfluxDBClient(host, port, user, password)
m = 'TAG01.10.11.22'
ms = all_measurements_lookup(client)
print ("400: ", ms[:10])
# ms_lijst = all_measurements_to_list(ms, 'measurementList.csv')
dbList = measurement_database_lookup(ms, m, debug='n')
print('500: ', f'Measurement {m} in database {dbList["database"]}.')
db = dbList['database'][0]
# print(db)

#
now = datetime.datetime.now()

mList = []
#print(ws[3])
for tag in ws[3][startColumnTags:]:  # start from first tag in Column 'startColumnTags'
    # print(tag, tag.value)
    if tag.value:
        mList.append(tag.value)
    else:
        break  # eliminate 'None' cells
# mList = mList[startColumnTags:]  # clean mList from non-tags
print('length mList, mList: ', len(mList), mList)


begin_time_valid = True    # True means: begin_date or end_date is valid date of type 01/01/1970 00:00:00,00 or later
end_time_valid = True      # True means: begin_date or end_date is valid date of type 01/01/1970 00:00:00,00 or later
# Get values
rij_nummer = 3
for row in list(ws.rows)[3:]:
    rij_nummer += 1
    print('rij_nummer:', rij_nummer)

    time_error_message = ''
    if row[0].value == None or isinstance(row[0].value, datetime.datetime):
        begin_time_valid = True
        time_error_message += " Begin time is valid;"
        #print(time_error_message)
    else:
        begin_time_valid = False
        time_error_message += " Begin time is NOT valid;"
        #print(time_error_message)
        row[3].value = time_error_message
        #break
    #print(row[0].value, isinstance(row[0].value, datetime.datetime))
    if row[1].value == 'now' or isinstance(row[1].value, datetime.datetime):
        end_time_valid = True
        time_error_message += " End time is valid;"
        #print(time_error_message)
    else:
        end_time_valid = False
        time_error_message += " End time is NOT valid;"
        #print(time_error_message)
        row[3].value = time_error_message
        #break
    begin_end_time_valid = begin_time_valid and end_time_valid


    print(row[0].value, row[1].value, row[2].value, datetime.datetime.now())
    if not row[0].value and row[1].value == 'now' and begin_end_time_valid:
        row[2].value = datetime.datetime.now()  # insert time of query in column 'C'

        timeE = datetime.datetime.now().isoformat()
        # timeE = datetime.datetime.strftime((timeE), '%Y-%m-%dT%H:%M:%S')
        print('timeE: ', timeE, type(timeE))
        timeE_Epoch = time_ISO8601_to_Influx_epoch(timeE)
        print('timeE: ', timeE_Epoch)

        nr = 0
        for m in mList:
            dbList = measurement_database_lookup(ms, m, debug='n')
            if len(dbList['database']) == 1:
                db = dbList['database'][0]
                print('db: ', db)
                client.switch_database(db)
                queryResult = []
                queryResult = client.query(
                    'select last(*) from \"{0}\" where time < {1}'.format(m, timeE_Epoch))
                queryResultList = list(queryResult.get_points())
                # print(queryResult)
                print(queryResultList)
            else:
                queryResultList = []
            if queryResultList:  # meaning if queryResultList is not empty
                print(len(queryResultList))
                print('m, nr, last_value: ', m, nr, queryResultList[0]['last_value'])
                row[startColumnTags + nr].value = queryResultList[0]['last_value']
            nr += 1


    # print(row[0].value, row[1].value, row[2].value, datetime.datetime.now())
    elif (not row[2].value or row[1].value != 'now') and begin_end_time_valid:
        if not row[0].value and row[1].value and not row[2].value and datetime.datetime.timestamp(now) >= datetime.datetime.timestamp(row[1].value):
            timeE = row[1].value
            timeE = datetime.datetime.strftime((timeE), '%Y-%m-%dT%H:%M:%S')
            print('timeE: ', timeE, type(timeE))
            timeE_Epoch = time_ISO8601_to_Influx_epoch(timeE)
            print('timeE: ', timeE_Epoch)

            row[2].value = datetime.datetime.now()  # insert time of query in column 'C'

            nr = 0
            for m in mList:
                dbList = measurement_database_lookup(ms, m, debug='n')
                if len(dbList['database']) == 1:
                    db = dbList['database'][0]
                    print('db: ', db)
                    client.switch_database(db)
                    queryResult = []
                    queryResult = client.query(
                        'select last(*) from \"{0}\" where time < {1}'.format(m, timeE_Epoch))
                    queryResultList = list(queryResult.get_points())
                    # print(queryResult)
                    print(queryResultList)
                else:
                    queryResultList = []
                if queryResultList:  # meaning if queryResultList is not empty
                    print(len(queryResultList))
                    print('m, nr, last_value: ', m, nr, queryResultList[0]['last_value'])
                    row[startColumnTags + nr].value = queryResultList[0][
                        'last_value']
                nr += 1

        if begin_end_time_valid and row[0].value and row[1].value and not row[2].value and datetime.datetime.timestamp(now) >= datetime.datetime.timestamp(row[1].value):
            print('now - row[2].value < 0: ', row[2].value)
            print(datetime.datetime.timestamp(now) - datetime.datetime.timestamp(row[1].value))

            timeB = row[0].value
            timeB = datetime.datetime.strftime((timeB), '%Y-%m-%dT%H:%M:%S')
            print(timeB, type(timeB))
            timeB_Epoch = time_ISO8601_to_Influx_epoch(timeB)
            print(timeB_Epoch)

            timeE = row[1].value
            timeE = datetime.datetime.strftime((timeE), '%Y-%m-%dT%H:%M:%S')
            print('timeE: ', timeE, type(timeE))
            timeE_Epoch = time_ISO8601_to_Influx_epoch(timeE)
            print('timeE: ',timeE_Epoch)

            row[2].value = datetime.datetime.now()  # insert time of query in column 'C'

            nr = 0
            for m in mList:
                dbList = measurement_database_lookup(ms, m, debug='n')
                if len(dbList['database']) == 1:
                    db = dbList['database'][0]
                    print('db: ', db)
                    client.switch_database(db)
                    queryResult = []
                    queryResult = client.query('select mean(*) from \"{0}\" where time > {1} and time <= {2}'.format(m, timeB_Epoch, timeE_Epoch))
                    queryResultList = list(queryResult.get_points())
                    # print(queryResult)
                    print(queryResultList)
                else:
                    queryResultList = []
                if queryResultList:  # meaning if queryResultList is not empty
                    print(len(queryResultList))
                    print('m, nr, mean_value: ', m, nr, queryResultList[0]['mean_value'])
                    row[startColumnTags+nr].value = queryResultList[0]['mean_value']
                nr += 1
    else:
        pass
        #break


wb.save(xlsxDirFileNameOut)

# query from 'list' : not in influxDB v0.9
# queryResultSet = client.query('select mean(*) from \'TAG1.10.11.22\',\'TAG2.20.22.33\' where time > {1} and time <= {2}'.format((timeB_Epoch, timeE_Epoch))) # not in v0.9

endRunTime = datetime.datetime.now()
print('endRunTime: ', endRunTime)
totalRunTime = endRunTime - startRunTime

print('\n\
Program run time statistics:\n\
{0:>30}: program start \n\
{1:>30}: program end \n\
{2:>30}: Total run time for this program'.format(str(startRunTime), str(endRunTime), str(totalRunTime))
)



